import openpyxl
import os
import re

wb_path = r"c:\Projects\theatre\Copy of Neiruz Tickets 2026.xlsm"
seed_sql_path = r"c:\Projects\theatre\db\seed.sql"

months_map = {
    "يناير": "01", "فبراير": "02", "مارس": "03", "ابريل": "04",
    "مايو": "05", "يونيو": "06", "يوليو": "07", "أغسطس": "08",
    "سبتمبر": "09", "أكتوبر": "10", "نوفمبر": "11", "ديسمبر": "12"
}

def parse_arabic_date(date_str):
    # Format: "الثلاثاء 1 سبتمبر 2026"
    if not date_str:
        return "2026-09-01"
    parts = date_str.split()
    if len(parts) >= 4:
        day = parts[1]
        month_name = parts[2]
        year = parts[3]
        
        # Normalize month name
        month = "09"
        for k, v in months_map.items():
            if k in month_name:
                month = v
                break
        
        day_int = int(day)
        return f"{year}-{month}-{day_int:02d}"
    return "2026-09-01"

if not os.path.exists(wb_path):
    print("Excel file not found!")
    exit(1)

try:
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheet = wb["Shows"]
    
    sql_lines = [
        "-- Seed Data imported from Copy of Neiruz Tickets 2026.xlsm\n",
        "BEGIN;\n\n",
        "TRUNCATE TABLE shows CASCADE;\n\n"
    ]
    
    # Row 1 is header
    # Row 2 to 32 contain show definitions
    for r in range(2, 33):
        row_id = sheet.cell(row=r, column=2).value
        if not row_id:
            continue
            
        date_arabic = sheet.cell(row=r, column=3).value
        # Parse arabic date to YYYY-MM-DD
        date_iso = parse_arabic_date(date_arabic)
        
        # Show description
        desc = sheet.cell(row=r, column=4).value or ""
        # Time
        raw_time = str(sheet.cell(row=r, column=5).value or "6:30م")
        if "8:00" in raw_time or "8" in raw_time:
            time_start = "20:00"
            time_end = "22:00"
        else:
            time_start = "18:30"
            time_end = "20:00"
        
        # Capacities
        cap_a = 125
        cap_b = 20
        cap_c = 110
        cap_d = 155
        
        # Clean desc
        desc_clean = desc.replace("'", "''")
        
        # Write show insert
        show_prefix = str(10 + int(row_id))
        sql_lines.append(f"-- Show {row_id}\n")
        sql_lines.append(
            f"INSERT INTO shows (id, name, prefix, date, time, end_time) VALUES ({row_id}, '{desc_clean}', '{show_prefix}', '{date_iso}', '{time_start}', '{time_end}');\n"
        )
        # Write zones inserts
        sql_lines.append(
            f"INSERT INTO zones (show_id, zone_name, total_capacity, available_seats) VALUES ({row_id}, 'A', {cap_a}, {cap_a});\n"
        )
        sql_lines.append(
            f"INSERT INTO zones (show_id, zone_name, total_capacity, available_seats) VALUES ({row_id}, 'B', {cap_b}, {cap_b});\n"
        )
        sql_lines.append(
            f"INSERT INTO zones (show_id, zone_name, total_capacity, available_seats) VALUES ({row_id}, 'C', {cap_c}, {cap_c});\n"
        )
        sql_lines.append(
            f"INSERT INTO zones (show_id, zone_name, total_capacity, available_seats) VALUES ({row_id}, 'D', {cap_d}, {cap_d});\n"
        )
        sql_lines.append("\n")
        
    sql_lines.append("COMMIT;\n")
    
    with open(seed_sql_path, "w", encoding="utf-8") as f:
        f.writelines(sql_lines)
    print("Seed SQL created successfully at", seed_sql_path)
    
except Exception as e:
    print("Error creating seed script:", e)
